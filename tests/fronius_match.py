"""
fronius_match.py
----------------
Busca todos os nomes de usinas retornados pela API Fronius Solar.web e gera um
arquivo 'fronius_nomes.xlsx' com os nomes na coluna A, em ordem alfabética.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from dotenv import load_dotenv

from modules.Fronius import buscar_dados_completos

load_dotenv(dotenv_path=Path(__file__).parent / ".env")

SEP   = "─" * 50
SAIDA = "fronius_nomes.xlsx"


def main() -> None:
    print(f"\n{SEP}")
    print("  FRONIUS MATCH — Exportação de nomes da API")
    print(SEP)

    print("\n  Conectando ao Fronius Solar.web...")
    todas_usinas = buscar_dados_completos()

    if not todas_usinas:
        print("  [ERRO] Nenhuma usina retornada pela API. Verifique as credenciais no .env.")
        return

    nomes_api = sorted(
        (u.get("name", "") for u in todas_usinas if u.get("name")),
        key=str.lower,
    )
    print(f"  {len(nomes_api)} usinas carregadas da API.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usinas Fronius"

    cabecalho = ws.cell(row=1, column=1, value="Nome da Usina (API Fronius)")
    cabecalho.font      = Font(bold=True, color="FFFFFF")
    cabecalho.fill      = PatternFill("solid", fgColor="FF6600")
    cabecalho.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 45

    for i, nome in enumerate(nomes_api, start=2):
        ws.cell(row=i, column=1, value=nome)

    try:
        wb.save(SAIDA)
        print(f"\n  Arquivo '{SAIDA}' salvo com {len(nomes_api)} nomes na coluna A.")
    except PermissionError:
        print(f"  [ERRO] Feche o arquivo '{SAIDA}' antes de rodar o script.")
        return

    print(f"{SEP}\n")


if __name__ == "__main__":
    main()
