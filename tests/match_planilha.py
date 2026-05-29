import csv
import openpyxl
from openpyxl.styles import PatternFill
from pathlib import Path

pasta_atual = Path(__file__).parent
caminho_planilha = pasta_atual / 'clientes.xlsx'
caminho_csv = pasta_atual / 'matches.csv'
caminho_nova_planilha = pasta_atual / 'clientes_atualizado.xlsx'

def atualizar_nomes():
    print("1. Lendo o arquivo matches.csv...")
    dicionario_matches = {}
    
    try:
        # utf-8-sig resolve problemas de caracteres invisíveis do Bloco de Notas do Windows
        with open(caminho_csv, mode='r', encoding='utf-8-sig') as f:
            # Descobre se o separador é vírgula ou ponto-e-vírgula
            conteudo = f.read()
            f.seek(0)
            delimitador = ';' if ';' in conteudo and ',' not in conteudo else ','
            
            leitor = csv.reader(f, delimiter=delimitador)
            next(leitor, None)  # Pula o cabeçalho
            
            for linha in leitor:
                if len(linha) >= 2:
                    # Remove espaços e coloca TUDO EM MAIÚSCULO para a comparação ser perfeita
                    nome_antigo = str(linha[0]).strip().upper()
                    nome_novo = str(linha[1]).strip()
                    dicionario_matches[nome_antigo] = nome_novo
    except Exception as e:
        print(f"❌ Erro ao ler CSV: {e}")
        return

    print(f"   -> 🔎 {len(dicionario_matches)} nomes carregados do dicionário!")
    
    if len(dicionario_matches) == 0:
        print("🛑 PARE: O script não conseguiu ler os nomes do matches.csv. Verifique o arquivo.")
        return

    print("2. Abrindo a planilha clientes.xlsx...")
    wb = openpyxl.load_workbook(caminho_planilha)
    aba_principal = wb['Planilha1'] 
    
    cor_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    cor_vermelha = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    alteracoes = 0
    
    print("3. Procurando os pares na planilha...")
    for row in range(2, aba_principal.max_row + 1):
        celula = aba_principal.cell(row=row, column=1)
        nome_original = str(celula.value).strip() if celula.value else ""
        
        if nome_original:
            # Deixa o nome do Excel em MAIÚSCULO também, só para comparar
            nome_para_comparar = nome_original.upper()
            
            if nome_para_comparar in dicionario_matches:
                # Achou! Coloca o nome oficial da Fronius/Growatt e pinta de verde
                celula.value = dicionario_matches[nome_para_comparar]
                celula.fill = cor_verde
                alteracoes += 1
            else:
                # Não achou. Pinta de vermelho
                celula.fill = cor_vermelha

    wb.save(caminho_nova_planilha)
    print(f"\n✅ Sucesso! {alteracoes} nomes foram atualizados e pintados de verde.")

if __name__ == "__main__":
    atualizar_nomes()