import openpyxl
from modules.Solis import buscar_dados_completos as solis_buscar, traduzir_status as solis_status
from modules.Growatt import buscar_dados_completos as growatt_buscar, traduzir_status as growatt_status
from modules.Fronius import buscar_dados_completos as fronius_buscar, traduzir_status as fronius_status
from modules.Solplanet import buscar_dados_completos as solplanet_buscar, traduzir_status as solplanet_status
from modules.Hypontech import buscar_dados_completos as hypontech_buscar, traduzir_status as hypontech_status
SEP = "─" * 50


def _log(mensagem):
    print(f"  {mensagem}")


def main():
    arquivo = 'clientes.xlsx'

    # ── Abertura da planilha ──────────────────────────────────
    print(f"\n{SEP}")
    print("  SINCRONIZADOR DE STATUS DE USINAS")
    print(SEP)

    try:
        wb = openpyxl.load_workbook(arquivo)
        sheet = wb.active
        _log(f"Planilha '{arquivo}' carregada.")
    except Exception as e:
        _log(f"[ERRO] Não foi possível abrir '{arquivo}': {e}")
        return

    # ── Detecção de marcas presentes ─────────────────────────
    marcas_presentes = set()
    for row in range(2, sheet.max_row + 1):
        marca = sheet.cell(row=row, column=6).value
        if marca:
            marcas_presentes.add(str(marca).strip().lower())

    # ── Sincronização com as APIs ─────────────────────────────
    print(f"\n{SEP}")
    print("  ETAPA 1 — Sincronização com APIs")
    print(SEP)

    base_solis = []
    base_growatt = []
    base_fronius = []
    base_solplanet = []
    base_hypontech = []

    if "solis" in marcas_presentes:
        _log("Conectando ao SolisCloud...")
        base_solis = solis_buscar()
        _log(f"Solis: {len(base_solis)} usinas carregadas.")

    if "fronius" in marcas_presentes:
        _log("Conectando ao Fronius Solar.web (status + geração por usina)...")
        base_fronius = fronius_buscar()
        _log(f"Fronius: {len(base_fronius)} usinas prontas (status, etoday, e_month).")

    if "solplanet" in marcas_presentes:
        _log("Conectando ao Solplanet...")
        base_solplanet = solplanet_buscar()
        _log(f"Solplanet: {len(base_solplanet)} usinas carregadas.")

    if "hypontech" in marcas_presentes:
        _log("Conectando ao Hypontech...")
        base_hypontech = hypontech_buscar()
        _log(f"Hypontech: {len(base_hypontech)} usinas carregadas.")

    # ── Processamento da planilha ─────────────────────────────
    print(f"\n{SEP}")
    print("  ETAPA 2 — Processamento da planilha")
    print(SEP)

    # Garante cabeçalhos
    if not sheet.cell(row=1, column=8).value:
        sheet.cell(row=1, column=8).value = "Última Atualização"
    if not sheet.cell(row=1, column=9).value:
        sheet.cell(row=1, column=9).value = "Geração Hoje (kWh)"
    if not sheet.cell(row=1, column=10).value:
        sheet.cell(row=1, column=10).value = "Geração Mensal (kWh)"

    total_linhas = sheet.max_row - 1
    processados = 0

    for row in range(2, sheet.max_row + 1):
        nome  = sheet.cell(row=row, column=1).value
        marca = str(sheet.cell(row=row, column=6).value or "").strip().lower()

        if not nome or not marca:
            continue

        if marca == "solis":
            status, last_update, etoday, e_month = solis_status(nome, base_solis)
            sheet.cell(row=row, column=7).value  = status
            sheet.cell(row=row, column=8).value  = last_update
            sheet.cell(row=row, column=9).value  = etoday
            sheet.cell(row=row, column=10).value = e_month
            processados += 1
            continue
        elif marca == "fronius":
            status, last_update, etoday, e_month = fronius_status(nome, base_fronius)
            sheet.cell(row=row, column=7).value  = status
            sheet.cell(row=row, column=8).value  = last_update
            sheet.cell(row=row, column=9).value  = etoday
            sheet.cell(row=row, column=10).value = e_month
            processados += 1
            continue
        elif marca == "solplanet":
            status, last_update, etoday, e_month = solplanet_status(nome, base_solplanet)
            sheet.cell(row=row, column=7).value = status
            sheet.cell(row=row, column=8).value = last_update
            sheet.cell(row=row, column=9).value = etoday
            sheet.cell(row=row, column=10).value = e_month
            processados += 1
            continue
        elif marca == "hypontech":
            status, last_update, etoday, e_month = hypontech_status(nome, base_hypontech)
            sheet.cell(row=row, column=7).value  = status
            sheet.cell(row=row, column=8).value  = last_update
            sheet.cell(row=row, column=9).value  = etoday
            sheet.cell(row=row, column=10).value = e_month
            processados += 1
            continue
        else:
            status, last_update = "Marca não mapeada", ""

        sheet.cell(row=row, column=7).value = status
        sheet.cell(row=row, column=8).value = last_update
        processados += 1

    _log(f"{processados}/{total_linhas} clientes processados.")

    # ── Salvamento ────────────────────────────────────────────
    print(f"\n{SEP}")
    print("  ETAPA 3 — Salvamento")
    print(SEP)

    try:
        wb.save(arquivo)
        _log(f"Arquivo '{arquivo}' salvo com sucesso.")
    except PermissionError:
        _log(f"[ERRO] Feche o arquivo '{arquivo}' antes de rodar o script.")

    print(f"\n{SEP}\n")


if __name__ == "__main__":
    main()
