"""
growatt_match.py
----------------
Gera uma aba "Growatt_Match" no clientes.xlsx com dois blocos em ordem
alfabética lado a lado:

  Bloco esquerdo  → nomes dos clientes marcados como "growatt" na planilha
  Bloco direito   → todos os nomes de usinas retornados pela API Growatt

O objetivo é facilitar a revisão manual: o usuário compara as duas colunas
e corrige os nomes na planilha principal conforme necessário.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from dotenv import load_dotenv

from modules.Growatt import buscar_dados_completos

load_dotenv(dotenv_path=Path(__file__).parent / ".env")

SEP    = "─" * 50
ARQUIVO = "clientes.xlsx"

COL_NOME  = 1  # coluna A — nome do cliente na aba principal
COL_MARCA = 6  # coluna F — marca do inversor


def _cabecalho(cell, texto):
    cell.value = texto
    cell.font  = Font(bold=True, color="FFFFFF")
    cell.fill  = PatternFill("solid", fgColor="2E75B6")
    cell.alignment = Alignment(horizontal="center")


def main() -> None:
    print(f"\n{SEP}")
    print("  GROWATT MATCH — Exportação para revisão manual")
    print(SEP)

    # ── Busca todas as usinas via API ─────────────────────────────────────────
    print("\n  Conectando ao Growatt...")
    todas_usinas = buscar_dados_completos()

    if not todas_usinas:
        print("  [ERRO] Nenhuma usina retornada pela API. Verifique o token no .env.")
        return

    nomes_api = sorted(
        [(u["name"], u.get("plant_id", "")) for u in todas_usinas],
        key=lambda x: x[0].lower(),
    )
    print(f"  {len(nomes_api)} usinas carregadas da API.")

    # ── Leitura da planilha ───────────────────────────────────────────────────
    print(f"\n  Lendo '{ARQUIVO}'...")
    try:
        wb = openpyxl.load_workbook(ARQUIVO)
    except Exception as e:
        print(f"  [ERRO] Não foi possível abrir '{ARQUIVO}': {e}")
        return

    sheet_principal = wb.active

    nomes_planilha = []
    for row in range(2, sheet_principal.max_row + 1):
        nome  = sheet_principal.cell(row=row, column=COL_NOME).value
        marca = sheet_principal.cell(row=row, column=COL_MARCA).value
        if nome and str(marca).strip().lower() == "growatt":
            nomes_planilha.append(str(nome).strip())

    nomes_planilha.sort(key=str.lower)
    print(f"  {len(nomes_planilha)} clientes Growatt encontrados na planilha.")

    # ── Cria/substitui aba de match ───────────────────────────────────────────
    ABA = "Growatt_Match"
    if ABA in wb.sheetnames:
        del wb[ABA]

    ws = wb.create_sheet(ABA)

    # Cabeçalhos
    _cabecalho(ws.cell(row=1, column=1), "Cliente (Planilha)")
    _cabecalho(ws.cell(row=1, column=3), "Usina (API Growatt)")
    _cabecalho(ws.cell(row=1, column=4), "plant_id")

    # Coluna separadora visual
    ws.column_dimensions["B"].width = 4

    # Larguras
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 20

    # Dados — cada bloco independente, ambos em ordem alfabética
    max_linhas = max(len(nomes_planilha), len(nomes_api))

    for i in range(max_linhas):
        linha = i + 2

        if i < len(nomes_planilha):
            ws.cell(row=linha, column=1).value = nomes_planilha[i]

        if i < len(nomes_api):
            nome_api, plant_id = nomes_api[i]
            ws.cell(row=linha, column=3).value = nome_api
            ws.cell(row=linha, column=4).value = plant_id

    # ── Salvar ────────────────────────────────────────────────────────────────
    print(f"\n  Salvando '{ARQUIVO}'...")
    try:
        wb.save(ARQUIVO)
        print(f"  Arquivo salvo. Abra a aba '{ABA}' para revisar os nomes.")
    except PermissionError:
        print(f"  [ERRO] Feche o arquivo '{ARQUIVO}' antes de rodar o script.")
        return

    # ── Relatório ─────────────────────────────────────────────────────────────
    print(f"\n{SEP}")
    print(f"  Clientes Growatt na planilha : {len(nomes_planilha)}")
    print(f"  Usinas retornadas pela API   : {len(nomes_api)}")
    print(f"  Aba gerada                   : {ABA}")
    print(f"{SEP}\n")


if __name__ == "__main__":
    main()
